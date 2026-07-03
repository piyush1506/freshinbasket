import openpyxl
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.http import HttpResponse
from django.utils.text import slugify
from store.models import Product, Category, Unit
import logging

logger = logging.getLogger('api')


class IsAdminRole(permissions.BasePermission):
    def has_permission(self, request, view):
        return (
            request.user
            and request.user.is_authenticated
            and request.user.role == 'ADMIN'
        )


class ProductTemplateDownloadView(APIView):
    """Download a sample Excel template for product import."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Header row
        headers = [
            'name', 'description', 'price', 'mrp', 'stock',
            'tax_percentage', 'category', 'unit'
        ]
        ws.append(headers)

        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 18

        # Sample rows
        samples = [
            ['Fresh Tomatoes', 'Farm fresh red tomatoes', 40, 50, 100, 0, 'Vegetables', 'kg'],
            ['Organic Spinach', 'Tender organic spinach leaves', 30, 40, 80, 0, 'Leafy Greens', 'bunch'],
            ['Alphonso Mangoes', 'Sweet Ratnagiri Alphonso mangoes', 250, 300, 50, 5, 'Fruits', 'kg'],
        ]
        for row in samples:
            ws.append(row)

        # Instructions sheet
        ws2 = wb.create_sheet("Instructions")
        ws2.append(["Column", "Required", "Description"])
        instructions = [
            ["name", "Yes", "Product name (must be unique)"],
            ["description", "No", "Product description"],
            ["price", "Yes", "Selling price (numeric)"],
            ["mrp", "No", "Maximum Retail Price (numeric)"],
            ["stock", "No", "Stock quantity (integer, default: 0)"],
            ["tax_percentage", "No", "Tax % e.g. 5, 12, 18 (default: 0)"],
            ["category", "No", "Category name (auto-created if not exists)"],
            ["unit", "No", "Unit name e.g. kg, bunch, piece (auto-created if not exists)"],
        ]
        for row in instructions:
            ws2.append(row)

        # Style instructions header
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            ws2.column_dimensions[cell.column_letter].width = 30

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
        return response


class ProductImportView(APIView):
    """Import products from an Excel file."""
    permission_classes = [IsAdminRole]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Only Excel files (.xlsx, .xls) are supported.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Invalid Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Read headers from row 1
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        required_cols = {'name', 'price'}
        missing = required_cols - set(headers)
        if missing:
            return Response(
                {'error': f'Missing required columns: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        def col(row_data, col_name):
            try:
                idx = headers.index(col_name)
                val = row_data[idx]
                return val if val is not None else ''
            except (ValueError, IndexError):
                return ''

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue  # Skip empty rows

            try:
                name = str(col(row, 'name')).strip()
                if not name or name.lower() == 'none':
                    errors.append({'row': row_idx, 'error': 'Name is required'})
                    skipped_count += 1
                    continue

                try:
                    price = float(col(row, 'price') or 0)
                except (ValueError, TypeError):
                    errors.append({'row': row_idx, 'error': f'Invalid price for "{name}"'})
                    skipped_count += 1
                    continue

                # Optional fields
                description = str(col(row, 'description') or '').strip()

                try:
                    mrp = float(col(row, 'mrp') or 0) or None
                except (ValueError, TypeError):
                    mrp = None

                try:
                    stock = int(float(col(row, 'stock') or 0))
                except (ValueError, TypeError):
                    stock = 0

                try:
                    tax_pct = float(col(row, 'tax_percentage') or 0)
                except (ValueError, TypeError):
                    tax_pct = 0

                # Category (auto-create)
                category_name = str(col(row, 'category') or '').strip()
                category = None
                if category_name and category_name.lower() != 'none':
                    category, _ = Category.objects.get_or_create(
                        name__iexact=category_name,
                        defaults={
                            'name': category_name,
                            'slug': slugify(category_name)
                        }
                    )

                # Unit (auto-create)
                unit_name = str(col(row, 'unit') or '').strip()
                unit = None
                if unit_name and unit_name.lower() != 'none':
                    unit, _ = Unit.objects.get_or_create(
                        name__iexact=unit_name,
                        defaults={
                            'name': unit_name,
                            'slug': slugify(unit_name)
                        }
                    )

                # Create or update product
                product, created = Product.objects.update_or_create(
                    name__iexact=name,
                    defaults={
                        'name': name,
                        'description': description,
                        'price': price,
                        'mrp': mrp,
                        'stock': stock,
                        'tax_percentage': tax_pct,
                        'unit': unit,
                    }
                )

                if category:
                    product.categories.add(category)

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except Exception as e:
                errors.append({'row': row_idx, 'error': str(e)})
                skipped_count += 1

        logger.info(
            f'Product import: created={created_count}, updated={updated_count}, '
            f'skipped={skipped_count}, errors={len(errors)} by user={request.user}'
        )

        return Response({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': errors[:20],  # Return max 20 errors
            'message': f'Import complete: {created_count} created, {updated_count} updated, {skipped_count} skipped.'
        })

